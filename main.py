from flask import Flask, jsonify, send_from_directory, request
from openpyxl import load_workbook
from flask_cors import CORS

app = Flask(__name__, static_folder='static')
CORS(app)

# SKU mapping
SKUMAP = {
    "860009592568": "Post-Bio-GH",
    "860009592551": "Omega-Alg",
    "850016364982": "Edi-DR-BC-SML",
    "864178000275": "Edi-DR-BC-LRG",
    "850016364968": "TS-Edi-HJ-PB",
    "850016364876": "Edi-HJ-PB-SML",
    "850016364883": "Edi-HJ-PB-LRG",
    "850016364890": "Edi-HJ-PB-FAM",
    "850016364951": "TS-Edi-STRESS-PB",
    "850016364838": "Edi-STRESS-PB-SML",
    "850016364852": "Edi-STRESS-PB-LRG",
    "850016364869": "Edi-STRESS-PB-FAM",
    "850016364906": "Edi-DR-SP-SML",
    "850016364913": "Edi-DR-SP-LRG",
    "850016364944": "TS-Edi-STRESS-Pepp",
    "850016364845": "Edi-STRESS-Pepp-SML",
    "850016364821": "Edi-STRESS-Pepp-LRG",
    "860008203403": "100-DR-HO",
    "860008203410": "200-DR-HO",
    "860008203427": "500-DR-HO",
    "860008203434": "750-DR-HO",
    "860009592575": "150-Mini-Stress-HO",
    "860008203441": "300-SR-HO",
    "860008203458": "600-SR-HO",
    "860008203465": "300-HJR-HO",
    "860008203472": "600-HJR-HO",
    "860008221988": "180-CAT-SR",
    "860008876775": "100-Lipe-Ultra",
    "860008876768": "300-Lipe-Ultra",
    "860009592513": "600-Lipe-Ultra",
    "861109000304": "CAP450",
    "850016364586": "SNT30",
    "860009592537": "TS-Itchy-Dry-Shampoo",
    "860008876713": "Itchy & Dry-SK-CT",
    "860009592520": "Itchy-Dry-Shampoo-Gallon",
    "860008876720": "Sensitive-SK-CT",
    "860008876737": "Conditioner-SK-CT",
    "860009592544": "TS-2in1-Shampoo",
    "860008876744": "2in1-SK-CT",
    "860008221971": "SK-PW-RL",
}

@app.route("/", methods=["GET"])
def home():
    return send_from_directory(app.static_folder, "index.html")

@app.route('/get-lot-codes', methods=['GET'])
def get_lot_codes():
    try:
        wb = load_workbook('LotCode.xlsx')
        ws = wb.active
        
        print("\n=== Starting SKU Processing ===")
        lot_codes = {}
        sku_columns = []
        current_sku = None
        
        # Find SKU columns
        header_row = next(ws.rows)
        for col_idx, cell in enumerate(header_row):
            if cell.value and 'SKU' in str(cell.value).upper():
                sku_columns.append(col_idx)
                print(f"Found SKU column at index {col_idx}")
        
        # Process each row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx in sku_columns:
                cell_value = row[col_idx].value
                if cell_value:
                    cell_value = str(cell_value).strip()
                    
                    # If we find a SKU (non-empty and not 'Total')
                    if cell_value and cell_value != 'Total':
                        current_sku = cell_value
                        if current_sku not in lot_codes:
                            lot_codes[current_sku] = {}
                            print(f"Processing new SKU: {current_sku}")
                        
                        # Now loop through subsequent rows until we hit 'Total'
                        current_row_idx = row_idx
                        while current_row_idx <= ws.max_row:
                            current_row = list(ws.rows)[current_row_idx - 1]
                            lot_value = current_row[col_idx + 1].value  # Lot number column
                            bb_date = current_row[col_idx + 4].value    # BB date column
                            
                            # Break if we hit 'Total' or empty cell
                            if not lot_value or str(lot_value).strip() == 'Total':
                                print(f"Hit Total or empty cell for {current_sku}")
                                break
                            
                            # Add lot code and BB date
                            if bb_date:
                                bb_date_str = bb_date.strftime('%m/%d/%y')
                            else:
                                bb_date_str = ''
                                
                            lot_codes[current_sku][str(lot_value)] = bb_date_str
                            print(f"Added lot code for {current_sku}: {lot_value} -> {bb_date_str}")
                            
                            current_row_idx += 1
        
        print("\n=== Final Lot Codes ===")
        for sku, lots in lot_codes.items():
            print(f"\n{sku}:")
            for lot, bb in lots.items():
                print(f"  {lot}: {bb}")
            
        return jsonify({"status": "success", "data": lot_codes})
    except Exception as e:
        print("Error loading lot codes:", str(e))
        return jsonify({"status": "error", "message": str(e)})

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)

