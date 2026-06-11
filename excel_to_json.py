import json
import os
import shutil
import tempfile
from datetime import datetime, timedelta

from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\GabbyEsquibel\Pet Releaf\Warehouse - Documents\NEW WH Locations 2026.xlsx"
SHEET_NAME = "Master Location List"
SKU_COL = 9   # column I
LOT_COL = 13  # column M
BB_COL = 14   # column N
START_ROW = 2


def format_bb_date(bb):
    if not bb:
        return ""

    if hasattr(bb, "strftime"):
        return bb.strftime("%Y-%m-%d")

    if isinstance(bb, (int, float)):
        try:
            excel_epoch = datetime(1900, 1, 1)
            date_obj = excel_epoch + timedelta(days=bb - 2)
            return date_obj.strftime("%Y-%m-%d")
        except Exception:
            return str(bb)

    bb_str = str(bb)
    if " 00:00:00" in bb_str:
        bb_str = bb_str.replace(" 00:00:00", "")
    return bb_str


def open_workbook(excel_path):
    """Open Excel file; copy to temp if the original is locked (e.g. open in Excel)."""
    try:
        return load_workbook(excel_path, data_only=True)
    except PermissionError:
        print(f"File locked, copying to temp: {excel_path}")
        tmp = os.path.join(tempfile.gettempdir(), "NEW_WH_Locations_2026_copy.xlsx")
        shutil.copy2(excel_path, tmp)
        return load_workbook(tmp, data_only=True)


def process_master_location_list(ws, lot_codes):
    """Read SKU, Lot #, and BB Date rows from Master Location List."""
    for row in range(START_ROW, ws.max_row + 1):
        sku = ws.cell(row=row, column=SKU_COL).value
        lot = ws.cell(row=row, column=LOT_COL).value

        if not sku or not str(sku).strip():
            continue
        if not lot or not str(lot).strip():
            continue

        sku = str(sku).strip()
        lot = str(lot).strip()

        if sku.lower() in ("", "total", "sku") or lot.lower() in ("", "total", "lot #"):
            continue

        bb_str = format_bb_date(ws.cell(row=row, column=BB_COL).value)

        if sku not in lot_codes:
            lot_codes[sku] = {}

        existing = lot_codes[sku].get(lot, {}).get("bb_date", "")
        if lot in lot_codes[sku] and not bb_str and existing:
            continue

        lot_codes[sku][lot] = {"bb_date": bb_str}
        print(f"  {sku} / {lot} -> BB {bb_str or '(empty)'}")


def load_existing_json():
    json_paths = ["static/js/lot_codes.json", "public/js/lot_codes.json"]
    existing_data = {}

    for path in json_paths:
        if os.path.exists(path):
            try:
                with open(path, "r") as f:
                    existing_data = json.load(f)
                print(f"Loaded existing data from {path}")
                break
            except (json.JSONDecodeError, FileNotFoundError):
                print(f"Could not load existing data from {path}")

    return existing_data


def compare_and_print_new_lots(existing_data, new_data):
    print("\n" + "=" * 60)
    print("NEW LOTS DETECTED:")
    print("=" * 60)

    new_lots_found = False
    all_new_lots = []

    for sku, lots in new_data.items():
        if sku not in existing_data:
            print(f"\nNEW SKU: {sku}")
            for lot, lot_info in lots.items():
                bb_date = lot_info.get("bb_date", "") if isinstance(lot_info, dict) else str(lot_info)
                print(f"  New lot: {lot} (BB: {bb_date})")
                all_new_lots.append({
                    "SKU": sku,
                    "Lot Code": lot,
                    "Best By Date": bb_date,
                    "Type": "New SKU",
                })
            new_lots_found = True
        else:
            new_lots_for_sku = []
            for lot, lot_info in lots.items():
                if lot not in existing_data[sku]:
                    new_lots_for_sku.append((lot, lot_info))

            if new_lots_for_sku:
                print(f"\nNEW LOTS for existing SKU: {sku}")
                for lot, lot_info in new_lots_for_sku:
                    bb_date = lot_info.get("bb_date", "") if isinstance(lot_info, dict) else str(lot_info)
                    print(f"  New lot: {lot} (BB: {bb_date})")
                    all_new_lots.append({
                        "SKU": sku,
                        "Lot Code": lot,
                        "Best By Date": bb_date,
                        "Type": "New Lot",
                    })
                new_lots_found = True

    if not new_lots_found:
        print("No new lots detected - all data is up to date!")
    else:
        try:
            from openpyxl import Workbook

            desktop_path = r"C:\Users\GabbyEsquibel\OneDrive - Pet Releaf\Desktop\newLots.xlsx"

            try:
                wb = load_workbook(desktop_path)
                ws = wb.active
                print("Loaded existing newLots.xlsx file")
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws.title = "New Lots Detected"
                headers = ["SKU", "Lot Code", "Best By Date", "Type", "Detection Date"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                print("Created new newLots.xlsx file")

            next_row = ws.max_row + 1
            detection_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for lot_data in all_new_lots:
                ws.cell(row=next_row, column=1, value=lot_data["SKU"])
                ws.cell(row=next_row, column=2, value=lot_data["Lot Code"])
                ws.cell(row=next_row, column=3, value=lot_data["Best By Date"])
                ws.cell(row=next_row, column=4, value=lot_data["Type"])
                ws.cell(row=next_row, column=5, value=detection_date)
                next_row += 1

            wb.save(desktop_path)
            print(f"Added {len(all_new_lots)} new lots to existing file: {desktop_path}")
        except Exception as e:
            print(f"Could not update newLots.xlsx: {e}")

    print("=" * 60)


def convert_excel_to_json():
    print(f"Reading: {EXCEL_PATH}")
    print(f"Sheet: {SHEET_NAME}")

    wb = open_workbook(EXCEL_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    existing_data = load_existing_json()

    lot_codes = {}
    process_master_location_list(ws, lot_codes)

    print(f"\nFinished: {len(lot_codes)} SKUs, {sum(len(lots) for lots in lot_codes.values())} lots")

    compare_and_print_new_lots(existing_data, lot_codes)

    with open("static/js/lot_codes.json", "w") as f:
        json.dump(lot_codes, f, indent=2)
    with open("public/js/lot_codes.json", "w") as f:
        json.dump(lot_codes, f, indent=2)

    try:
        with open("public/index.html", "r") as f:
            content = f.read()
        if not content.endswith(" "):
            content += " "
        with open("public/index.html", "w") as f:
            f.write(content)
        print("Added deployment trigger to index.html")
    except Exception as e:
        print(f"Could not update index.html: {e}")

    print("\nSaved updated lot codes to JSON files")
    wb.close()


if __name__ == "__main__":
    convert_excel_to_json()
